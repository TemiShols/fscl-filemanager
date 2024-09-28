from langchain_community.document_loaders import Docx2txtLoader, PyPDFLoader, CSVLoader, SitemapLoader, YoutubeLoader, \
    UnstructuredURLLoader, SeleniumURLLoader
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from youtube_transcript_api import YouTubeTranscriptApi, TranscriptsDisabled, NoTranscriptFound, VideoUnavailable
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from langchain.schema import Document
import time
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urljoin, urlparse
import xml.etree.ElementTree as ET
from gtts import gTTS
import os
from .models import Project
import openpyxl


def load_docx_file(file_path):
    loader = Docx2txtLoader(file_path)
    return loader.load()


def load_pdf_file(file_path):
    loader = PyPDFLoader(file_path)
    return loader.load()


def load_csv_file(file_path):
    loader = CSVLoader(file_path)
    return loader.load()


def load_xlsx_file(file_path):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Initialize a list to store the content
        all_content = []

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read all rows and columns in the sheet
            for row in sheet.iter_rows(values_only=True):
                # Join each cell in the row into a single string and add to the list
                row_content = ' '.join([str(cell) for cell in row if cell is not None])
                all_content.append(row_content)

        # Join all rows together into one large string
        combined_content = '\n'.join(all_content)

        return combined_content

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


def load_sitemap_file(sitemap_url):
    loader = SitemapLoader(web_path=sitemap_url)
    return loader.load()


def load_youtube_file(url):
    try:
        if 'v=' in url:
            video_id = url.split('v=')[-1].split('&')[0]
        elif 'youtu.be/' in url:
            video_id = url.split('youtu.be/')[-1].split('?')[0]
        else:
            return "Invalid YouTube URL."

        # Fetch the transcript
        transcript = YouTubeTranscriptApi.get_transcript(video_id)
        full_text = " ".join([item['text'] for item in transcript])
        return full_text

    except TranscriptsDisabled:
        return "Transcripts are disabled for this video."
    except NoTranscriptFound:
        return "No transcript found for this video."
    except VideoUnavailable:
        return "The video is unavailable."
    except Exception as e:
        return f"An unexpected error occurred: {str(e)}"


def load_multiple_url(urls):
    url = [urls]
    loader = UnstructuredURLLoader(url, ssl_verify=False)
    return loader.load()


def load_url(urls):
    url = [urls]

    loader = SeleniumURLLoader(urls=url)

    return loader.load()


def loads_urls(urls):
    options = Options()
    options.headless = True
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("start-maximized")
    options.add_argument("enable-automation")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-browser-side-navigation")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/91.0.4472.124 Safari/537.36")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    all_text_content = []

    try:
        for url in urls:
            try:
                driver.get(url)
                # time.sleep(5)  # Let the page load completely
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                text_content = soup.get_text()
                all_text_content.append(text_content)
            except Exception as e:
                print(f"Error fetching URL {url}: {e}")

        # Join all text contents together
        full_text_content = "\n".join(all_text_content)
        return Document(metadata={"sources": urls}, page_content=full_text_content)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None
    finally:
        driver.quit()


def generate_sitemap(base_url):
    """Generates a sitemap for the given base URL."""
    options = Options()
    options.headless = True
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("start-maximized")
    options.add_argument("enable-automation")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-browser-side-navigation")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/91.0.4472.124 Safari/537.36"
    )

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    sitemap = ET.Element("urlset", xmlns="http://www.sitemaps.org/schemas/sitemap/0.9")
    visited_urls = set()

    def crawl(url):
        """Crawls a URL and discovers links."""
        if url in visited_urls or not url.startswith(base_url):
            return
        visited_urls.add(url)

        try:
            driver.get(url)
            time.sleep(5)  # Let the page load completely
            soup = BeautifulSoup(driver.page_source, 'html.parser')
        except Exception as e:
            print(f"Error crawling {url}: {e}")
            return

        url_element = ET.SubElement(sitemap, "url")
        ET.SubElement(url_element, "loc").text = url

        for link in soup.find_all("a", href=True):
            link_url = urljoin(url, link["href"])
            parsed_url = urlparse(link_url)
            normalized_url = parsed_url._replace(fragment="").geturl()  # Remove URL fragments
            crawl(normalized_url)

    try:
        crawl(base_url)
    finally:
        driver.quit()

    # Save the sitemap to the user's desktop securely
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "sitemap.xml")
    os.makedirs(os.path.dirname(desktop_path), exist_ok=True)  # Ensure the directory exists
    with open(desktop_path, "wb") as f:
        f.write(ET.tostring(sitemap, encoding="utf-8", method="xml"))

    return desktop_path


def extract_and_save_content(sitemap_xml_content, proj_id):
    """
    Extracts URLs from a sitemap XML content, fetches content, and saves it to a Django model using Selenium.

    Args:
        sitemap_xml_content (str): The XML content of the sitemap.
        proj_id (int): The ID of the project to save content to.
    """

    def extract_urls_from_sitemap(sitemap_xml_content):
        root = ET.fromstring(sitemap_xml_content)
        urls = []
        for url in root.findall('{http://www.sitemaps.org/schemas/sitemap/0.9}url'):
            loc = url.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc').text

            urls.append(loc)
        return urls

    # Set up Selenium WebDriver with Chrome options
    options = Options()
    options.headless = True
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("start-maximized")

    options.add_argument("enable-automation")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-browser-side-navigation")

    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-blink-features=AutomationControlled")

    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/91.0.4472.124 Safari/537.36")

    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        urls = extract_urls_from_sitemap(sitemap_xml_content)

        for url in urls:
            try:
                driver.get(url)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body")))  # Wait for page to load
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                text_content = soup.get_text()

                project = Project.objects.get(pk=proj_id)
                project.content = project.content + text_content if project.content else text_content
                project.save()
            except Exception as e:
                print(f"Error processing URL {url}: {e}")
    except Exception as e:
        print(f"Error setting up WebDriver: {e}")
    finally:
        if 'driver' in locals():
            driver.quit()


def text_to_speech(text):
    # Determine the desktop directory path
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")

    # Ensure the desktop directory exists (it should, by default)
    if not os.path.exists(desktop_dir):
        raise FileNotFoundError("Desktop directory not found")

    # Create a file name and full path for the audio file
    audio_file = f"{text[:10].replace(' ', '_')}.mp3"
    audio_file_path = os.path.join(desktop_dir, audio_file)

    # Convert text to speech and save it to the desktop
    tts = gTTS(text=text, lang='en')
    tts.save(audio_file_path)

    return audio_file_path
