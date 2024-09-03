from mistralai.client import MistralClient
from mistralai.models.chat_completion import ChatMessage
import numpy as np
import pandas as pd
import os, re
import faiss
from docx import Document
import openpyxl
import csv
import requests
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
from mistralai.client import MistralClient
from PyPDF2 import PdfReader
from django.conf import settings
import chromadb


chroma_client = chromadb.Client()

client = MistralClient(api_key=settings.MISTRAL_UTIL_API_KEY)


#   collection = Collection()


def clean_text(text):
    # Convert to lowercase
    text = text.lower()
    # Remove special characters
    text = re.sub(r'\W+', ' ', text)
    # Remove extra spaces
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def read_docx(file):
    if file.name.split('.')[-1] == 'docx':
        doc = Document(file)
        full_text = []
        for paragraph in doc.paragraphs:
            full_text.append(paragraph.text)
        return '\n'.join(full_text)


def read_xlxs(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data


def format_xlxs_for_prompt(content):
    output_text = ""
    for row in content:
        output_text += "\t".join([str(cell) for cell in row]) + "\n"
    return output_text


def save_xlsx_to_text(input_filename, text_output_filename):
    workbook = openpyxl.load_workbook(input_filename)
    with open(text_output_filename, 'w', newline='') as text_file:
        writer = csv.writer(text_file, delimiter='\t')
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            writer.writerow([f'Sheet: {sheet_name}'])
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)


def process_xlsx(file):
    df = pd.read_excel(file, engine='openpyxl')
    null_percentage = df.isnull().mean().mean()

    if null_percentage > 0.5:
        return 'Null values are too much'

    df.fillna('', inplace=True)
    text_output = df.applymap(str).agg(' '.join, axis=1).str.cat(sep=' ')
    return clean_text(text_output)


def read_sitemap(url):
    response = requests.get(url)
    if response.status_code == 200:
        sitemap = response.content
        root = ET.fromstring(sitemap)
        urls = []
        for url in root.findall('.//{http://www.sitemaps.org/schemas/sitemap/0.9}url'):
            loc = url.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc').text
            urls.append(loc)
        return urls
    else:
        print(f"Failed to fetch sitemap: {response.status_code}")
        return []


def read_pdf(file):
    pdf_reader = PdfReader(file)
    text = ''
    for page in pdf_reader.pages:
        text += page.extract_text()
    return text


def generate_sitemap(base_url):
    """Generates a sitemap for the given base URL."""
    sitemap = ET.Element("urlset", xmlns="http://www.sitemaps.org/schemas/sitemap/0.9")

    # Crawl the base URL to discover links
    response = requests.get(base_url)
    soup = BeautifulSoup(response.content, "html.parser")

    for link in soup.find_all("a", href=True):
        url = link["href"]
        if url.startswith(base_url):
            url_element = ET.SubElement(sitemap, "url")
            ET.SubElement(url_element, "loc").text = url

    # Write the sitemap to a file
    with open("sitemap.xml", "wb") as f:
        f.write(ET.tostring(sitemap))


# Example usage
#   base_url = "https://www.example.com"
#   generate_sitemap(base_url)

#   print("Sitemap generated successfully!")


def scrape_content(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        # You can customize the scraping logic depending on the structure of the web pages
        # Example: Extracting all paragraphs
        paragraphs = soup.find_all('p')
        content = '\n'.join([para.get_text() for para in paragraphs])
        return content
    else:
        print(f"Failed to fetch URL: {url} - Status code: {response.status_code}")
        return ""


def scrape_sitemap(sitemap_url):
    urls = read_sitemap(sitemap_url)
    all_content = {}
    for url in urls:
        print(f"Scraping {url}...")
        content = scrape_content(url)
        all_content[url] = content
    return all_content


def get_text_embedding(input):
    embeddings_batch_response = client.embeddings(
        model="mistral-embed",
        input=input
    )
    return embeddings_batch_response.data[0].embedding


def chunk_texts(file, chunk_size=512):
    return [file[i:i + chunk_size] for i in range(0, len(file), chunk_size)]


def construct_index(chunks):
    text_embeddings = np.array([get_text_embedding(chunk) for chunk in chunks])
    dimension = text_embeddings.shape[1]
    index = faiss.IndexFlatL2(dimension)
    print(index.ntotal)
    index.add(text_embeddings)
    return index


def construct_index_chromadb(chunks):
    # Create a new collection
    collection = chroma_client.get_or_create_collection(name="fileapp")

    # Get text embeddings
    text_embeddings = [get_text_embedding(chunk) for chunk in chunks]

    # Add text embeddings to the collection
    collection.add(
        ids=[str(i) for i in range(len(chunks))],
        embeddings=text_embeddings,
        metadatas=None
    )

    return collection


def run_mistral(user_message, model="mistral-medium-latest"):
    messages = [
        ChatMessage(role="user", content=user_message)
    ]
    chat_response = client.chat(
        model=model,
        messages=messages
    )
    return chat_response.choices[0].message.content


def rag_mistral(query, collection, chunks, k=2, model="mistral-medium-latest"):
    query_embeddings = get_text_embedding(query)
    results = collection.query(query_embeddings, n_results=k)

    print(f"Query: {query}")
    print(f"Results: {results}")

    if not results or 'ids' not in results or not results['ids'] or not results['ids'][0]:
        return "No results found"

    retrieved_chunk_ids = results['ids'][0]
    retrieved_chunks = [chunks[int(i)] for i in retrieved_chunk_ids if i.isdigit()]

    if not retrieved_chunks:
        return "No relevant information found"

    context = ' '.join(retrieved_chunks)
    prompt = f"""
            Context information is below.
            ---------------------
            {context}
            ---------------------
            Given the context information and not prior knowledge, answer the query.
            Query: {query}
            Answer:
            """
    return run_mistral(prompt, model=model)


def summarize_content(content, max_tokens=500):
    prompt = f""" Please provide a concise summary of the following text. Focus on the main points and key information
        ---------------------
        {content}
        ---------------------
        Summary:"""
    summary = run_mistral(prompt)
    return summary


def analyze_content(content, max_tokens=500):
    prompt = f"""
    Please analyze the following spreadsheet content and provide a concise summary. Focus on the main data points, trends, and any notable insights or patterns.

    Spreadsheet Content:
    ---------------------
    {content}
    ---------------------

    Summary:
    """
    analysis = run_mistral(prompt)
    return analysis
